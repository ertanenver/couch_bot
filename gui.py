import sys
import time
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

# Путь к директории с файлом
DIRECTORY_PATH = "/home/shinyaa/my_projects/couch_bot"

# Имя файла
FILE_NAME = 'excel.xlsx'

# Индексы столбцов в файле с данными
NAME_COLUMN_INDEX = 0
RATING_COLUMN_INDEX = 1


class RatingTableApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Rating Table App")

        # Создание виджетов
        self.table_widget = QTableWidget()
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.table_widget)
        self.setLayout(self.layout)

        # Заполнение таблицы данными
        self.update_table()

        # Отслеживание изменений в директории
        self.observer = Observer()
        self.observer.schedule(FileModifiedEventHandler(self.handle_file_changed), path=DIRECTORY_PATH, recursive=False)
        self.observer.start()

    def update_table(self):
        # Чтение данных из файла Excel и расчет рейтинга
        wb = load_workbook(filename=f"{DIRECTORY_PATH}/{FILE_NAME}")
        ws = wb.active

        # Очистка таблицы
        self.table_widget.clear()

        # Задание количества строк и столбцов таблицы
        self.table_widget.setRowCount(ws.max_row)
        self.table_widget.setColumnCount(ws.max_column)

        # Заполнение таблицы данными
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
            name = row[NAME_COLUMN_INDEX]
            rating = row[RATING_COLUMN_INDEX]
            self.table_widget.setItem(i, 0, QTableWidgetItem(str(name)))
            self.table_widget.setItem(i, 1, QTableWidgetItem(str(rating)))

    def handle_file_changed(self, event):
        # Обработчик события изменения файла
        if event.src_path.endswith('.xlsx'):
            self.update_table()

    def closeEvent(self, event):
        self.observer.stop()
        self.observer.join()
        event.accept()


class FileModifiedEventHandler(FileSystemEventHandler):
    """Обработчик событий изменения файла"""
    def __init__(self, callback):
        super().__init__()
        self.callback = callback

    def on_modified(self, event):
        self.callback(event)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    rating_table_app = RatingTableApp()
    sys.exit(app.exec_())

