import os
import configparser

from reportlab.lib.styles import getSampleStyleSheet

styles = getSampleStyleSheet()
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
#from PIL import ImageTk
import qrcode
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl
#from rotatedtext import verticalText

from tkinter import *
from tkinter import filedialog, simpledialog, ttk
import tkinter as tk
pdfmetrics.registerFont (TTFont('Nash', "calibri.ttf"))
tk = []
ccc = True
nasnach1 = 'Благотворительный взнос за '




def settings():
    def save_config():
        global ent
        global nasnach1
        config = configparser.ConfigParser()
        config.add_section("Purpose")
        # config['Purpose'] = {'Назначение: ': ent.get()}
        config['Purpose'] = {'User_input': ent.get()}
        with open(r'SETTINGS.ini', 'w') as configfile:
            config.write(configfile)
        nasnach1 = config.get("Purpose", "user_input")
        root1.destroy()


        return nasnach1
    global ent
    global nasnach1
    root1 = Tk()
    tk.append(root1)
    root1.title("Настройки")
    root1.geometry("440x150")
   # root1.eval('tk::PlaceWindow . center')

    label1 = ttk.Label(tk[1], text="Назначение платежа")
    label1.pack(expand=True)

    def fetch():
        print('Ввод = "%s"' % ent.get())  # извлечь строку
        #ent.delete(0, END)  # очистить поле ввода

    ent = Entry(root1)
    ent.insert(0, 'Благотворительный взнос за ')  # записать строку начиная с 0 позиции в поле ввода
    ent.pack(side=TOP, fill=X)
    ent.focus()  # передать фокус в поле ввода
    ent.bind('<Return>', (lambda event: fetch()))  # завершение ввода клавишей Enter
    Button(root1, text='OK', command=save_config).pack(side=BOTTOM)



def read_excel():
    global label
    global btn
    global nasnach1
    gruppa=""
    FileExcel = filedialog.askopenfilename(
                                          title="Загружаемый файл Excel",
                                          filetypes=(("Файл Excel",
                                                      "*.xlsx*"),
                                                     ("all files",
                                                      "*.*")))
    FilePDF = filedialog.asksaveasfilename(defaultextension=".pdf",
                                          title="Название сохраняемого файла PDF",
                                          filetypes=(("Файл PDF",
                                                      "*.pdf*"),
                                                     ("all files",
                                                      "*.*")))

    c = canvas.Canvas(FilePDF, pagesize=A4)
    if nasnach1 != None:
        naznach = nasnach1 + simpledialog.askstring("Назначение платежа", nasnach1)
    else:
        naznach = 'Благотворительный взнос за ' + simpledialog.askstring("Назначение платежа", 'Благотворительный взнос за ')



    # создаем окно "Ожидайте..."
    label.configure(text="Ожидайте...")
    label.update()

    wb = openpyxl.load_workbook(FileExcel)
    for ws in wb:
        F = ''
        I = ''
        O = ''

        a = 1
        aa=1

        while ws.cell(a, 1).value != None:
            Summ = ws.cell(a , 2).value
            if Summ == None or isinstance(Summ, str):
                Summ = 0

            FIO_TAB = ws.cell(a, 1).value
            if FIO_TAB != None:
                FIO_TAB = str(ws.cell(a, 1).value)
            if FIO_TAB == None:
                FIO_TAB = ''
            FIO_LIST = FIO_TAB.split()
            try:
                    F = FIO_LIST[0]
            except:
                    F = ''

            try:
                    I = FIO_LIST[1]
            except:
                    I = ''

            try:
                    O = FIO_LIST[2]
            except:
                    O = ''
            print(F, I, O, str(Summ))

            if isinstance(ws.cell(a, 2).value, int):
                if Summ > 0:
                    if (aa + 3) / 4 == (aa + 3) // 4:
                        if aa > 4:
                            c.showPage()
                        plateghka(628, c, a, F, I, O, Summ, naznach)
                        c.rotate(90)
                        c.drawString(10, -585, gruppa)
                        c.rotate(-90)

                       # c.drawString(10, 10, verticalText(gruppa))

                        aa += 1

                    else:
                        if (aa + 2) / 4 == (aa + 2) // 4:
                            plateghka(419, c, a, F, I, O, Summ, naznach)
                            aa += 1
                        else:
                            if (aa + 1) / 4 == (aa + 1) // 4:
                                plateghka(210, c, a, F, I, O, Summ, naznach)
                                aa += 1
                            else:
                                if (aa) / 4 == (aa) // 4:
                                    plateghka(1, c, a, F, I, O, Summ, naznach)
                                    aa += 1
            else:
                if a>2 and aa>1:
                    c.showPage()
                aa=1
                gruppa=ws.cell(a, 1).value

            a += 1
    c.save()
    label.configure(text="Файл готов")
    label.update()


def plateghka(y,c,a,F, I, O, Summ, naznach):
    text = 'ST00012|Name=РОО "Спортивная федерация Тхэквондо (МФТ) Тамбовской области"|PersonalAcc=40703810761000000482|BankName=В ОТДЕЛЕНИИ ТАМБОВ|BIC=046850649|CorrespAcc=00000000000000000000|Sum=' + str(Summ) + '00|Purpose=' + naznach + '|PayeeINN=6829152581|LastName=' + F + '|FirstName=' + I + '|MiddleName=' + O + '|TechCode=08'

    c.rect(10, y+4, 560, 205)
    c.rect(10, y+4, 120, 205)
    c.rect(130, y+158, 440, 17)
    c.rect(130, y+101, 440, 30)
    c.rect(130, y+63, 440, 15)
    c.rect(130, y+25, 440, 23)

    img = qrcode.make(text)
    print(text)
    type(img)
    img.save("znach"+str(a)+".png")
    c.drawImage('znach'+str(a)+'.png',13,y+54,114,114)

    c.setFont('Nash', 10)
    c.drawString(44, y + 184, "Извещение")
    c.drawString(485, y + 199, "Форма №пд-4")
    c.drawString(135, y + 199, "СБЕРБАНК РОССИИ")
    c.drawString(135, y + 179, 'РОО "Спортивная федерация Тхэквондо (МФТ) Тамбовской области"')
    c.drawString(140, y + 162, "6829152581")
    c.drawString(230, y + 162, "КПП")
    c.drawString(330, y + 162, "№ 40703810761000000482")
    c.drawString(135, y + 135, "в В ОТДЕЛЕНИИ ТАМБОВ")
    c.drawString(285, y + 135, "БИК 046850649")
    c.drawString(135, y + 116, "ОКТМО")
    c.drawString(195, y + 116, "КБК")
    c.drawString(135, y + 105, naznach)
    c.drawString(325, y + 105, "р.н.")
    c.drawString(135, y + 81, "Ф.И.О. плательщика")
    c.drawString(235, y + 81, F + " " + I + " " + O)
    c.drawString(135, y + 67, "Адрес плательщика")
    c.drawString(135, y + 52, "ИНН плательщика")
    c.drawString(135, y + 29, "Итого")
    c.drawString(135, y + 39, "Сумма платежа")
    c.drawString(295, y + 39, "Сумма платы за услуги    0 руб. 00 коп.")
    c.drawString(295, y + 29, "Дата")
    c.drawString(210, y + 39, str(Summ) + " руб. 00 коп.")
    c.drawString(210, y + 29, str(Summ) + " руб. 00 коп.")
    c.setFont('Nash', 7)
    c.drawString(135, y + 19, "С условиями приёма указанной в платежном документе суммы, в т.ч. с суммой взимаемой платы за услуги банка, ознакомлен и согласен")
    c.drawString(135, y + 8, "Подпись плательщика ________________")
    c.drawString(270, y + 170, "(наименование получателя платежа)")
    c.drawString(135, y + 151, "(ИНН получателя платежа)")
    c.drawString(330, y + 150, "(номер счета получателя платежа)")
    c.drawString(180, y + 126, "(наименование банка получателя платежа)")
    c.drawString(170, y + 94, "(наименование платежа)")
    c.drawString(300, y + 93, "(рег.номер плательщика)")
    c.drawString(430, y + 92, "(код принадлежности)")
    os.remove("znach"+str(a)+".png")


if __name__ == '__main__':
    root = Tk()
    tk.append(root)
    root.title("Квитанция МФТ")
    #root.geometry("1920x1080")
    root.geometry("440x150")
   # root.eval('tk::PlaceWindow . center')

   # image = ImageTk.PhotoImage(file="ima.png")
    Button(text='Настройки', command=settings).pack(anchor="ne", padx=5, pady=5)
    label = ttk.Label(tk[0], text="Нажмите старт")
    label.pack(expand=True)
    btn = Button(text="Старт", command=read_excel)  # создаем кнопку из пакета ttk
    btn.pack(expand=True)  # размещаем кнопку в окне

    root.mainloop()