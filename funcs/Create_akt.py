from openpyxl.styles import PatternFill
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment

# couch type: {"fio":["kindergarden 1","schoool 35"]}
months = {'01': 'январь', '02': 'февраль', '03': 'март', '04': 'апрель', '05': 'май', '06': 'июнь', '07': 'июль', '08': 'август', '09': 'сентябрь', '10': 'октябрь', '11': 'ноябрь', '12': 'декабрь'}

def create_akt(couch:dict,code='09'):
    for key, value in couch.items():
        arr = value
        fio = key
    wb = Workbook()
    ws = [wb.create_sheet(title=x) for x in arr]
    del wb['Sheet']
    today = datetime.now().strftime("%d.%m.%Y")
    
    for el in wb.sheetnames:
        ws = wb[el]
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['C'].width = 13
        ws.freeze_panes = "B3"
        ws.merge_cells('A1:A2')
        ws['A1'] = "ФИО"
        ws.merge_cells('B1:C1')  
        ws['B1'] = "Общее"
        ws['B2'] = "Долг"
        ws['C2'] = "Переплата"
        ws.merge_cells("D1:F1")
        ws["D1"] = months['09'].title()
        ws['D2'] = 'опл'
        ws["E2"] = 'дата'
        ws['F2'] = 'квит'


        if int(code)  != 9:
            ws.merge_cells("G1:I1")
            ws["G1"] = months['10'].title()
            ws['G2'] = 'опл'
            ws["H2"] = 'дата'
            ws['I2'] = 'квит'
            if int(code) != 10:
                ws.merge_cells("J1:L1")
                ws["J1"] = months['11'].title()
                ws['J2'] = 'опл'
                ws["K2"] = 'дата'
                ws['L2'] = 'квит'
                if int(code) != 11:
                    ws.merge_cells("M1:O1")
                    ws["M1"] = months['12'].title()
                    ws['M2'] = 'опл'
                    ws["N2"] = 'дата'
                    ws['O2'] = 'квит'
                    if int(code) != 12:
                        ws.merge_cells("P1:R1")
                        ws["P1"] = months['01'].title()
                        ws['P2'] = 'опл'
                        ws["Q2"] = 'дата'
                        ws['R2'] = 'квит'
                        if int(code) != 1:
                            ws.merge_cells("S1:U1")
                            ws["S1"] = months['02'].title()
                            ws['S2'] = 'опл'
                            ws["T2"] = 'дата'
                            ws['U2'] = 'квит'
                            if int(code) != 2:
                                ws.merge_cells("V1:X1")
                                ws["V1"] = months['03'].title()
                                ws['V2'] = 'опл'
                                ws["W2"] = 'дата'
                                ws['X2'] = 'квит'
                                if int(code) != 3:
                                    ws.merge_cells("Y1:AA1")
                                    ws["Y1"] = months['04'].title()
                                    ws['Y2'] = 'опл'
                                    ws["Z2"] = 'дата'
                                    ws['AA2'] = 'квит'
                                    if int(code) != 4:
                                        ws.merge_cells("AB1:AD1")
                                        ws["AB1"] = months['05'].title()
                                        ws['AB2'] = 'опл'
                                        ws["AC2"] = 'дата'
                                        ws['AD2'] = 'квит'
                                        if int(code) != 5:
                                            ws.merge_cells("AE1:AG1")
                                            ws["AE1"] = months['06'].title()
                                            ws['AE2'] = 'опл'
                                            ws["AF2"] = 'дата'
                                            ws['AG2'] = 'квит'
                                            if int(code) != 6:
                                                ws.merge_cells("AH1:AJ1")
                                                ws["AH1"] = months['07'].title()
                                                ws['AH2'] = 'опл'
                                                ws["AI2"] = 'дата'
                                                ws['AJ2'] = 'квит'
                                                if int(code) != 7:
                                                    ws.merge_cells("AK1:AM1")
                                                    ws["AK1"] = months['08'].title()
                                                    ws['AK2'] = 'опл'
                                                    ws["AL2"] = 'дата'
                                                    ws['AM2'] = 'квит'




            

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Выравнивание по центру ячейки и вертикально по центру
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        

    # Сохраняем файл
    wb.save(f'Акт сверки {fio} на {today}.xlsx')




create_akt({'Шиняев Владимир Олегович':["Дс 70", "Медвежонок","Золотой ключик"]}, code='09')