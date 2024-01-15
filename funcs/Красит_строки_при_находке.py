from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# Загружаем файл
workbook = load_workbook('example.xlsx')

# Получаем активный лист
sheet = workbook.active

# Текст для поиска
text = "Пелагеевские"

# Цвет для раскраски
fill_color = 'FFFF00' # Желтый цвет

# Создаем стиль для раскраски
fill = PatternFill(fill_type='solid', fgColor=fill_color)

# Перебираем все строки в листе
for row in sheet.iter_rows():
    # Перебираем все ячейки в строке
    for cell in row:
        # Проверяем содержимое ячейки на частичное совпадение с текстом
        if cell.value and re.search(r'\b{}\b'.format(text), str(cell.value), re.IGNORECASE):
            # Получаем адрес ячейки
            cell_address = cell.coordinate
            # Применяем стиль к ячейкам всей строки
            for c in row:
                c.fill = fill
            # Добавляем текст после найденной информации в ячейке
            cell.value = cell.value + " - найдено '{}'".format(text)
            # Выходим из цикла, чтобы не обрабатывать одну строку несколько раз
            break

# Сохраняем изменения в файле
workbook.save('example.xlsx')