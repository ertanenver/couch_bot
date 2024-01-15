import openpyxl

def coordinate(Excel:str):
    if '.xlsx' not in Excel:
        return 'Неправильный формат файла'

    wb = openpyxl.load_workbook(Excel)
    for ws in wb:
        # Список значений для поиска
        values = ['id', 'id_tg','fio','nickname','phone_num','tmp_password']

        cells = {}

        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value in values:
                    cells[cell.value] = cell.coordinate

    return cells

def excel_couch(Excel:str):
    cells = coordinate(Excel)
    wb = openpyxl.load_workbook(Excel)
    array = []

    for ws in wb:
        b = ws.min_row
        dict = {}
        while b != ws.max_row:
            id_tg_col = ws[cells['id_tg']].column
            fio_col = ws[cells['fio']].column
            nickname_col = ws[cells['nickname']].column
            phone_num_col = ws[cells['phone_num']].column
            tmp_password_col = ws[cells['tmp_password']].column

            id_tg = ws.cell(b, id_tg_col).value
            fio = ws.cell(b, fio_col).value
            nickname = ws.cell(b, nickname_col).value
            phone_num = ws.cell(b, phone_num_col).value
            tmp_password = ws.cell(b, tmp_password_col).value
            array = [fio, nickname, phone_num, tmp_password, id_tg]
            dict[fio] = array

    return dict
