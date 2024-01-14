import os

from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

import qrcode
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl

styles = getSampleStyleSheet()

pdfmetrics.registerFont (TTFont('calibri', "calibri.ttf"))


def read_excel(FileExcel, FilePDF = 'Квитанции.pdf', purpose = 'Благотворительный взнос за ', month = '', feature = ''):

    gruppa=""

    c = canvas.Canvas(FilePDF, pagesize=A4)
    feature = " " + feature if feature != "" else ""
    purpose += month + feature

    wb = openpyxl.load_workbook(FileExcel)
    for ws in wb:
        F = ''
        I = ''
        O = ''

        number_kvit_on_page = 1
        column_index = ws.min_row
        
        while column_index - 1 < ws.max_row:
            if ws.cell(column_index , 1).value == None:
                pass
            else:
                Summ = ws.cell(column_index , 2).value
                if Summ == None or isinstance(Summ, str):
                    Summ = 0
                FIO_TAB = ws.cell(column_index, 1).value
                if FIO_TAB != None:
                    FIO_TAB = str(FIO_TAB)
                else:
                    FIO_TAB = ''
                FIO_LIST = FIO_TAB.split()
                try:
                        F = FIO_LIST[0]
                except:
                        F = ''
                try:
                        I = FIO_LIST[1]
                except:
                        I = ''
                try:
                        O = FIO_LIST[2]
                except:
                        O = ''
                fio = FIO_TAB
                if isinstance(ws.cell(column_index, 2).value, int):
                    if Summ > 0:
                        if (number_kvit_on_page + 3) / 4 == (number_kvit_on_page + 3) // 4:
                            if number_kvit_on_page > 4:
                                c.showPage()
                            plateghka(628, c, column_index, F, I, O, Summ, purpose, fio)
                            c.rotate(90)
                            c.drawString(10, -585, gruppa)
                            c.rotate(-90)
                            number_kvit_on_page += 1
                        else:
                            if (number_kvit_on_page + 2) / 4 == (number_kvit_on_page + 2) // 4:
                                plateghka(419, c, column_index, F, I, O, Summ, purpose, fio)
                                number_kvit_on_page += 1
                            else:
                                if (number_kvit_on_page + 1) / 4 == (number_kvit_on_page + 1) // 4:
                                    plateghka(210, c, column_index, F, I, O, Summ, purpose, fio)
                                    number_kvit_on_page += 1
                                else:
                                    if (number_kvit_on_page) / 4 == (number_kvit_on_page) // 4:
                                        plateghka(1, c, column_index, F, I, O, Summ, purpose, fio)
                                        number_kvit_on_page += 1
                else:
                    if column_index > 2 and number_kvit_on_page > 1:
                        c.showPage()
                    number_kvit_on_page = 1
                    gruppa=ws.cell(column_index, 1).value
            
        
            column_index += 1
    c.save()


def plateghka(y, c, column_index, F, I, O, Summ, purpose, fio):
    text = 'ST00012|Name=РОО "Спортивная федерация Тхэквондо (МФТ) Тамбовской области"|PersonalAcc=40703810761000000482|BankName=В ОТДЕЛЕНИИ ТАМБОВ|BIC=046850649|CorrespAcc=00000000000000000000|Sum=' + str(Summ) + '00|Purpose=' + purpose + " / " +F + " " + I + " " + O + '|PayeeINN=6829152581|LastName=' + F + '|FirstName=' + I + '|MiddleName=' + O + '|TechCode=08'

    c.rect(10, y+4, 560, 205)
    c.rect(10, y+4, 120, 205)
    c.rect(130, y+158, 440, 17)
    c.rect(130, y+101, 440, 30)
    c.rect(130, y+63, 440, 15)
    c.rect(130, y+25, 440, 23)

    img = qrcode.make(text)
    img.save("znach"+str(column_index)+".png")
    c.drawImage('znach'+str(column_index)+'.png',13,y+54,114,114)

    c.setFont('calibri', 10)
    c.drawString(44, y + 184, "Извещение")
    c.drawString(485, y + 199, "Форма №пд-4")
    c.drawString(135, y + 199, "СБЕРБАНК РОССИИ")
    c.drawString(135, y + 179, 'РОО "Спортивная федерация Тхэквондо (МФТ) Тамбовской области"')
    c.drawString(140, y + 162, "6829152581")
    c.drawString(230, y + 162, "КПП")
    c.drawString(330, y + 162, "№ 40703810761000000482")
    c.drawString(135, y + 135, "в В ОТДЕЛЕНИИ ТАМБОВ")
    c.drawString(285, y + 135, "БИК 046850649")
    c.drawString(135, y + 116, "ОКТМО")
    c.drawString(195, y + 116, "КБК")
    c.drawString(135, y + 105, purpose + " / " + fio + "      р.н.")
    #c.drawString(325, y + 105, "р.н.")
    c.drawString(135, y + 81, "Ф.И.О. плательщика")
    c.drawString(235, y + 81, F + " " + I + " " + O)
    c.drawString(135, y + 67, "Адрес плательщика")
    c.drawString(135, y + 52, "ИНН плательщика")
    c.drawString(135, y + 29, "Итого")
    c.drawString(135, y + 39, "Сумма платежа")
    c.drawString(295, y + 39, "Сумма платы за услуги    0 руб. 00 коп.")
    c.drawString(295, y + 29, "Дата")
    c.drawString(210, y + 39, str(Summ) + " руб. 00 коп.")
    c.drawString(210, y + 29, str(Summ) + " руб. 00 коп.")
    c.setFont('calibri', 7)
    c.drawString(135, y + 19, "С условиями приёма указанной в платежном документе суммы, в т.ч. с суммой взимаемой платы за услуги банка, ознакомлен и согласен")
    c.drawString(135, y + 8, "Подпись плательщика ________________")
    c.drawString(270, y + 170, "(наименование получателя платежа)")
    c.drawString(135, y + 151, "(ИНН получателя платежа)")
    c.drawString(330, y + 150, "(номер счета получателя платежа)")
    c.drawString(180, y + 126, "(наименование банка получателя платежа)")
    c.drawString(170, y + 94, "(наименование платежа)")
    c.drawString(300, y + 93, "(рег.номер плательщика)")
    c.drawString(430, y + 92, "(код принадлежности)")
    os.remove("znach"+str(column_index)+".png")


