import openpyxl

# Открываем файл Excel
workbook = openpyxl.load_workbook('example.xlsx')

# Получаем список всех листов в файле
sheets = workbook.sheetnames

# Проходим по каждому листу и раскрашиваем строки
for sheet_name in sheets:
    # Получаем лист по его имени
    sheet = workbook[sheet_name]

    # Инициализируем флаг, который указывает, нужно ли продолжать раскрашивать
    continue_coloring = True

    # Проходим по строкам и раскрашиваем их до ячейки A1 с надписью "ИТОГО"
    for row in sheet.iter_rows():
        # Если встретили ячейку A1 с надписью "ИТОГО", прерываем цикл
        if row[0].value == "ИТОГО":
            break

        # Раскрашиваем строку светло-зеленым цветом, если нужно продолжать раскрашивать
        if continue_coloring:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")



# Сохраняем изменения в файле Excel
workbook.save('example.xlsx')



#CCFFFF - Синий
#FFFFCC - Желтый
#ffc0cb - розовый

