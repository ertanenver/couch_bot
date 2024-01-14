import re
from openpyxl import Workbook
def toExcel(text:str, Excel:str):
    text = text.replace('\t',' ').replace('\n',' ')
    pattern = r'\b(\D+\d+)\b'
    elements = re.findall(pattern, text)

    wb = Workbook()
    ws = wb.active

    for row, item in enumerate(elements, start=1):
        fio, amount = item.rsplit(maxsplit=1)
        ws.cell(row=row, column=1, value=fio)
        ws.cell(row=row, column=2, value=int(amount))

    wb.save(Excel)

